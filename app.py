from flask import Flask, render_template, Response,jsonify,request
import cv2
import threading
from simple_facerec.simple_facerec import SimpleFacerec
from openpyxl import load_workbook
from datetime import datetime

# Load the existing Excel workbook
wb = load_workbook("C:/Users/91637/Downloads/attendance/detected_names.xlsx")
ws = wb.active

# Assuming your existing column headers are "Time stamp" and "Detected Names"
timestamp_column_header = "Time stamp"
column_header = "Detected Names"

app = Flask(__name__)

# Initialize SimpleFacerec and load face encodings
sfr = SimpleFacerec()
sfr.load_encoding_images("Images/")

# Lock for thread safety
lock = threading.Lock()

# Variables to store captured frame
captured_frame = None
detected_names = ""

def capture_faces():
    global captured_frame
    global detected_names

    cap = cv2.VideoCapture(0)
    
    while True:
        ret, frame = cap.read()
        face_locations, face_names = sfr.detect_known_faces(frame)

        for face_loc, name in zip(face_locations, face_names):
            y1, x2, y2, x1 = face_loc[0], face_loc[1], face_loc[2], face_loc[3]
            cv2.putText(frame, name, (x1, y1 - 10), cv2.FONT_HERSHEY_DUPLEX, 1, (0, 0, 0), 3)
            cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 0, 200), 3)
            detected_names = name 

        cv2.imshow("Frame", frame)
        key = cv2.waitKey(1)

        if key == 27:
           break

        if key == ord('c'):
           with lock:
               captured_frame = frame.copy()
               detected_names = detected_names.strip()

    cap.release()
    cv2.destroyAllWindows()


# # Start the face capture thread
# capture_thread = threading.Thread(target=capture_faces)
# capture_thread.daemon = True
# capture_thread.start()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/get_frame')
def get_frame():
    # Start the face capture thread
    capture_thread = threading.Thread(target=capture_faces)
    capture_thread.daemon = True
    capture_thread.start()
    global captured_frame

    

    if captured_frame is not None:
        _, jpeg = cv2.imencode('.jpg', captured_frame)
        return Response(b'--frame\r\n'
                        b'Content-Type: image/jpeg\r\n\r\n' + jpeg.tobytes() + b'\r\n\r\n',
                        mimetype='multipart/x-mixed-replace; boundary=frame')
    else:
        return "No frame available"

@app.route('/get_detected_names')
def get_detected_names():
    global detected_names
    # with lock:
    print(f"Detected Names: {detected_names}")  # Add this line to print the detected names
    return jsonify(detected_names.split('\n'))

##to save reg no in excel sheet
@app.route('/save_to_excel', methods=['POST'])
def save_to_excel():
    global detected_names

    data = request.json
    register_number = data.get('register_number', '')

    # Check if detected_names is not empty before saving
    if detected_names:
        with lock:
            # Find the column index of the existing column headers
            timestamp_column_index = 1  # Assuming it's the first column
            column_index = 2  # Assuming the second column is "Detected Names"
            for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=1, max_row=1):
                if col[0].value == timestamp_column_header:
                    timestamp_column_index = col[0].column
                elif col[0].value == column_header:
                    column_index = col[0].column

            # Find the next available row in the existing column
            next_row = ws.max_row + 1

            # Set the value of the next available cell in the timestamp column
            ws.cell(row=next_row, column=timestamp_column_index, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

            # Set the value of the next available cell in the "Detected Names" column
            ws.cell(row=next_row, column=column_index, value=detected_names.strip())

            # Save the modified Excel workbook
            wb.save("C:/Users/91637/Downloads/attendance/detected_names.xlsx")

        # Clear the detected_names after saving
        detected_names = ""

        return jsonify({'message': 'Data saved successfully'})
    else:
        return jsonify({'message': 'Detected names is empty. No data saved.'})


if __name__ == '__main__':
    app.run(debug=True)
